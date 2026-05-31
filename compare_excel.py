import pandas as pd
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def compare_excel_sheets(file_old, file_new, sheet_name, key_col):
    print(f"正在讀取並比對工作表: '{sheet_name}' (主鍵欄位: '{key_col}')...")
    
    # 1. Read files
    try:
        df_old = pd.read_excel(file_old, sheet_name=sheet_name)
        df_new = pd.read_excel(file_new, sheet_name=sheet_name)
    except Exception as e:
        print(f"錯誤: 無法讀取 Excel 檔案或工作表。詳細原因: {e}")
        return

    # 2. Data Cleaning - Strip whitespaces from string columns
    for df in [df_old, df_new]:
        # Clean column names
        df.columns = df.columns.str.strip()
        # Clean key column
        if key_col in df.columns:
            df[key_col] = df[key_col].astype(str).str.strip()
            # Remove any empty rows or rows where key is empty
            df.dropna(subset=[key_col], inplace=True)
        else:
            print(f"錯誤: 欄位 '{key_col}' 不存在於檔案中。")
            print(f"舊檔欄位: {list(df_old.columns)}")
            print(f"新檔欄位: {list(df_new.columns)}")
            return

    # Get all unique columns from both files (excluding the key column)
    all_cols = list(df_old.columns)
    for col in df_new.columns:
        if col not in all_cols:
            all_cols.append(col)
    
    other_cols = [c for c in all_cols if c != key_col]

    # Convert to dictionaries for fast lookup
    dict_old = df_old.set_index(key_col).to_dict(orient='index')
    dict_new = df_new.set_index(key_col).to_dict(orient='index')

    keys_old = set(dict_old.keys())
    keys_new = set(dict_new.keys())

    # Calculate differences
    added_keys = sorted(list(keys_new - keys_old))
    deleted_keys = sorted(list(keys_old - keys_new))
    common_keys = keys_old & keys_new

    modified_projects = []
    unchanged_projects = []

    for key in sorted(list(common_keys)):
        row_old = dict_old[key]
        row_new = dict_new[key]
        
        changes = {}
        for col in other_cols:
            val_old = row_old.get(col, None)
            val_new = row_new.get(col, None)
            
            # Treat NaN and None as same
            if pd.isna(val_old) and pd.isna(val_new):
                continue
            
            # For floats, avoid small float precision mismatch issues
            if isinstance(val_old, (int, float)) and isinstance(val_new, (int, float)):
                if abs(val_old - val_new) < 1e-6:
                    continue
            
            if val_old != val_new:
                changes[col] = {"old": val_old, "new": val_new}
        
        if changes:
            modified_projects.append({
                "key": key,
                "changes": changes,
                "row_old": row_old,
                "row_new": row_new
            })
        else:
            unchanged_projects.append(key)

    # Print Text Report to Console
    print("\n" + "="*50)
    print(" 📊 企業級專案比對結果摘要 (Real-world Corporate Summary)")
    print("="*50)
    print(f"🔹 舊檔專案總數: {len(keys_old)} 個")
    print(f"🔹 新檔專案總數: {len(keys_new)} 個")
    print(f"➕ 新增專案 (Added): {len(added_keys)} 個 {added_keys if added_keys else ''}")
    print(f"➖ 刪除專案 (Deleted): {len(deleted_keys)} 個 {deleted_keys if deleted_keys else ''}")
    print(f"⚠️ 變更專案 (Modified): {len(modified_projects)} 個")
    print(f"✅ 無變動專案: {len(unchanged_projects)} 個")
    print("="*50)

    if modified_projects:
        print("\n✏️ 變更細節:")
        for idx, item in enumerate(modified_projects, 1):
            row_new = item["row_new"]
            proj_name = row_new.get("專案名稱", "未知專案")
            print(f"  {idx}. 專案編號: 【{item['key']}】 - {proj_name}")
            for col, chg in item['changes'].items():
                old_val = chg['old']
                new_val = chg['new']
                # Format output values based on column type
                if "預算" in col:
                    old_val = f"${old_val:,.0f}"
                    new_val = f"${new_val:,.0f}"
                elif "進度" in col:
                    old_val = f"{old_val * 100:.1f}%"
                    new_val = f"{new_val * 100:.1f}%"
                print(f"     📍 欄位「{col}」變更: {old_val} ➡️ {new_val}")
        print("="*50)

    # ------------------ Generate Premium Styled Excel Report ------------------
    wb = openpyxl.Workbook()
    wb.remove(wb.active) # Remove default sheet

    # Define Styles
    font_family = "Microsoft JhengHei"
    
    title_font = Font(name=font_family, size=16, bold=True, color="1F497D")
    section_font = Font(name=font_family, size=12, bold=True, color="2C3E50")
    header_font = Font(name=font_family, size=11, bold=True, color="FFFFFF")
    data_font = Font(name=font_family, size=11, color="000000")
    bold_data_font = Font(name=font_family, size=11, bold=True, color="000000")
    
    # Beautiful Pastel Fills (Premium Design Aesthetics)
    fill_header = PatternFill(start_color="1F497D", end_color="1F497D", fill_type="solid") # Classic corporate dark navy
    fill_added = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")  # Elegant pastel green
    fill_deleted = PatternFill(start_color="FCE4D6", end_color="FCE4D6", fill_type="solid") # Elegant pastel coral
    fill_modified = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid") # Elegant pastel yellow
    fill_zebra = PatternFill(start_color="F9FAFB", end_color="F9FAFB", fill_type="solid")
    
    border_thin = Side(border_style="thin", color="D3D3D3")
    grid_border = Border(left=border_thin, right=border_thin, top=border_thin, bottom=border_thin)

    align_center = Alignment(horizontal="center", vertical="center", wrap_text=True)
    align_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    align_right = Alignment(horizontal="right", vertical="center", wrap_text=True)

    # Helper function to format values for cell write
    def write_formatted_cell(cell, value, col_name):
        cell.border = grid_border
        cell.font = data_font
        
        # Standard formats based on column names
        if "預算" in col_name and isinstance(value, (int, float)):
            cell.value = value
            cell.number_format = '"$"#,##0'
            cell.alignment = align_right
        elif "進度" in col_name and isinstance(value, (int, float)):
            cell.value = value
            cell.number_format = '0.0%'
            cell.alignment = align_right
        elif "日期" in col_name:
            cell.value = str(value)
            cell.alignment = align_center
        elif col_name in [key_col, "目前狀態", "負責部門", "專案經理 (PM)"]:
            cell.value = str(value)
            cell.alignment = align_center
        else:
            cell.value = str(value)
            cell.alignment = align_left

    # SHEET 1: Summary Sheet (比較總覽)
    ws_summary = wb.create_sheet(title="比較總覽")
    ws_summary.views.sheetView[0].showGridLines = True
    
    ws_summary["B2"] = "📊 企業級專案組合比對報告"
    ws_summary["B2"].font = title_font
    
    ws_summary["B4"] = "基本統計"
    ws_summary["B4"].font = section_font
    
    summary_headers = ["指標項目", "數量 (個)", "備註說明"]
    for col_num, header in enumerate(summary_headers, start=2):
        cell = ws_summary.cell(row=5, column=col_num, value=header)
        cell.font = header_font
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = grid_border
    
    summary_data = [
        ("原始專案數量 (舊檔)", len(keys_old), "比對前的專案組合清單總數"),
        ("最新專案數量 (新檔)", len(keys_new), "比對後的最新專案組合清單總數"),
        ("新增專案 (Added)", len(added_keys), "只存在於最新檔案的專案 (整列標記為淡綠色)"),
        ("刪除專案 (Deleted)", len(deleted_keys), "已在最新檔案中移除的專案 (整列標記為淡紅色)"),
        ("屬性變更專案 (Modified)", len(modified_projects), "專案存在但內容（如進度、預算、PM、狀態等）有調整 (單格標記為淡黃色)"),
        ("完全一致專案 (Unchanged)", len(unchanged_projects), "最新與最舊檔案中的內容完全吻合的專案")
    ]
    
    for row_idx, data in enumerate(summary_data, start=6):
        for col_idx, val in enumerate(data, start=2):
            cell = ws_summary.cell(row=row_idx, column=col_idx, value=val)
            cell.font = data_font
            cell.border = grid_border
            if col_idx == 2:
                cell.alignment = align_left
            elif col_idx == 3:
                cell.alignment = align_center
                cell.font = bold_data_font
            else:
                cell.alignment = align_left
                
            if data[0].startswith("新增") and col_idx == 3:
                cell.fill = fill_added
            elif data[0].startswith("刪除") and col_idx == 3:
                cell.fill = fill_deleted
            elif data[0].startswith("屬性變更") and col_idx == 3:
                cell.fill = fill_modified

    # SHEET 2: Full Comparison Sheet (完整比對明細表)
    ws_full = wb.create_sheet(title="完整比對明細表")
    ws_full.views.sheetView[0].showGridLines = True
    
    ws_full["A1"] = "📋 專案比對明細表 (淡綠=新增，淡紅=刪除，淡黃=欄位有變更)"
    ws_full["A1"].font = title_font
    
    headers = [key_col, "比對狀態"] + other_cols
    for col_num, header in enumerate(headers, start=1):
        cell = ws_full.cell(row=3, column=col_num, value=header)
        cell.font = header_font
        cell.fill = fill_header
        cell.alignment = align_center
        cell.border = grid_border
    
    current_row = 4
    
    # 1. Added rows (Green)
    for key in added_keys:
        row_data = dict_new[key]
        ws_full.cell(row=current_row, column=1, value=key).alignment = align_center
        ws_full.cell(row=current_row, column=2, value="➕ 新增").alignment = align_center
        
        ws_full.cell(row=current_row, column=1).font = data_font
        ws_full.cell(row=current_row, column=2).font = data_font
        ws_full.cell(row=current_row, column=1).border = grid_border
        ws_full.cell(row=current_row, column=2).border = grid_border
        
        for col_idx, col in enumerate(other_cols, start=3):
            cell = ws_full.cell(row=current_row, column=col_idx)
            write_formatted_cell(cell, row_data.get(col, ""), col)
            
        for c in range(1, len(headers) + 1):
            ws_full.cell(row=current_row, column=c).fill = fill_added
        current_row += 1

    # 2. Deleted rows (Red)
    for key in deleted_keys:
        row_data = dict_old[key]
        ws_full.cell(row=current_row, column=1, value=key).alignment = align_center
        ws_full.cell(row=current_row, column=2, value="➖ 刪除").alignment = align_center
        
        ws_full.cell(row=current_row, column=1).font = data_font
        ws_full.cell(row=current_row, column=2).font = data_font
        ws_full.cell(row=current_row, column=1).border = grid_border
        ws_full.cell(row=current_row, column=2).border = grid_border
        
        for col_idx, col in enumerate(other_cols, start=3):
            cell = ws_full.cell(row=current_row, column=col_idx)
            write_formatted_cell(cell, row_data.get(col, ""), col)
            
        for c in range(1, len(headers) + 1):
            ws_full.cell(row=current_row, column=c).fill = fill_deleted
        current_row += 1

    # 3. Modified rows (Yellow cell-level highlighting)
    for item in modified_projects:
        key = item["key"]
        row_old = item["row_old"]
        row_new = item["row_new"]
        changes = item["changes"]
        
        ws_full.cell(row=current_row, column=1, value=key).alignment = align_center
        ws_full.cell(row=current_row, column=2, value="⚠️ 變更").alignment = align_center
        
        ws_full.cell(row=current_row, column=1).font = data_font
        ws_full.cell(row=current_row, column=2).font = data_font
        ws_full.cell(row=current_row, column=1).border = grid_border
        ws_full.cell(row=current_row, column=2).border = grid_border
        
        for col_idx, col in enumerate(other_cols, start=3):
            cell = ws_full.cell(row=current_row, column=col_idx)
            
            if col in changes:
                old_val = changes[col]["old"]
                new_val = changes[col]["new"]
                
                # Dynamic value formatting in cell for change visibility
                if "預算" in col:
                    old_str = f"${old_val:,.0f}"
                    new_str = f"${new_val:,.0f}"
                elif "進度" in col:
                    old_str = f"{old_val * 100:.1f}%"
                    new_str = f"{new_val * 100:.1f}%"
                else:
                    old_str = str(old_val)
                    new_str = str(new_val)
                
                cell.value = f"【舊】{old_str}\n➡️ 【新】{new_str}"
                cell.alignment = Alignment(wrap_text=True, vertical="center", horizontal="left")
                cell.font = bold_data_font
                cell.fill = fill_modified
                cell.border = grid_border
            else:
                write_formatted_cell(cell, row_new.get(col, ""), col)
                
        current_row += 1

    # 4. Unchanged rows (Normal Zebra striped)
    for idx, key in enumerate(unchanged_projects):
        row_data = dict_new[key]
        ws_full.cell(row=current_row, column=1, value=key).alignment = align_center
        ws_full.cell(row=current_row, column=2, value="✅ 一致").alignment = align_center
        
        ws_full.cell(row=current_row, column=1).font = data_font
        ws_full.cell(row=current_row, column=2).font = data_font
        ws_full.cell(row=current_row, column=1).border = grid_border
        ws_full.cell(row=current_row, column=2).border = grid_border
        
        for col_idx, col in enumerate(other_cols, start=3):
            cell = ws_full.cell(row=current_row, column=col_idx)
            write_formatted_cell(cell, row_data.get(col, ""), col)
            
        for c in range(1, len(headers) + 1):
            cell = ws_full.cell(row=current_row, column=c)
            if idx % 2 == 1:
                cell.fill = fill_zebra
        current_row += 1

    # Autofit columns beautifully
    for col in ws_summary.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                val_str = str(cell.value)
                char_len = sum(2 if ord(char) > 256 else 1 for char in val_str)
                if char_len > max_len:
                    max_len = char_len
        ws_summary.column_dimensions[col_letter].width = min(max(max_len + 4, 12), 45)
        
    # Full sheet gets larger widths because of rich text wrap
    for col in ws_full.columns:
        col_letter = get_column_letter(col[0].column)
        col_name = ws_full.cell(row=3, column=col[0].column).value
        
        max_len = 0
        for cell in col[2:]: # Check starting from headers downwards
            if cell.value:
                # If cell has newline, measure the longest line
                lines = str(cell.value).split('\n')
                for line in lines:
                    char_len = sum(2 if ord(char) > 256 else 1 for char in line)
                    if char_len > max_len:
                        max_len = char_len
                        
        width = max(max_len + 4, 12)
        if col_name == "主要風險與備註說明" or col_name == "專案名稱":
            # Give long text columns a fixed, readable wrapping width
            width = 38
        ws_full.column_dimensions[col_letter].width = width

    # Save comparison report
    output_filename = "project_comparison_report.xlsx"
    wb.save(output_filename)
    print(f"\n✨ 完美生成『企業級專案對比報告』: '{output_filename}'，已根據商業格式優化了數值顯示與自動換行！")
    print("="*50)

if __name__ == "__main__":
    compare_excel_sheets(
        file_old="project_data_v1.xlsx",
        file_new="project_data_v2.xlsx",
        sheet_name="專案清單",
        key_col="專案編號"
    )
